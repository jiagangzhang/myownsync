import math
import sys
import itertools
from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl import load_workbook


to_file = 'split-product'
from_file='dataheavy_product_verify.xlsx'
file_count = 30

def count_exist_rows(work_sheet):
    count=0
    for row in work_sheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value and len(cell.value)>0:
                count+=1
    return count

# generate a list of A to Z......Z
def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)
        size +=1

#  create column from A,B,C to AL,AM... default is 39 columns(AM)
def generate_columnlist(opt=39):
    columnlist=[]
    for s in itertools.islice(iter_all_strings(), opt):
        columnlist.append(s)
    return columnlist

# cache the format of excel of the first row if delete all existing records
def cache_row(work_sheet,n=1, row_size=[]):
    cache=[]
    for s in row_size:
        cache.append(work_sheet['%s%d'%(s,n)].value)
    return cache

def main():
    original_file = from_file
    filename = to_file
    args = sys.argv[1:]
    if not args:
      print ("will split to %d copies, if you want to specify number of copies, use 'python3 xxx.py numberofcopies'"%file_count)
      howmanynewfiles=file_count
    else:
      try:
          howmanynewfiles=int(args[0])
          if howmanynewfiles<=0:
              print ('\n\tWrong format\tuse python3 xxx.py numberofcopies')
              sys.exit(1)
      except:
          print ("usage: [Number] [input File Name] [output File Name]. "+'Must specify an integer')
          sys.exit(1)

    del args[0:1]
    if not args:
        print ('will split file %s into %d copies named %s-n'%(original_file,file_count,filename))
    else:
        original_file=str(args[0])
        del args[0:1]
        if not args:
            print ('will split file %s into %d copies named %s-n'%(original_file,file_count,filename))
        else:
            filename=str(args[0])

    print ('will split file %s into %d copies named %s-n'%(original_file,howmanynewfiles,filename))

    try:
        wb=load_workbook(filename=original_file,data_only = True)
    except:
        print ("File not exists or corrupted, please try again")
        sys.exit(1)

    # wb=load_workbook(filename=original_file,read_only=True)
    ws=wb.active
    # print ws.title
    ws_new_title=ws.title



    exist_count=count_exist_rows(ws)
    print ('\n\tExist:',exist_count)
    how_many_rows_per_file = int(math.floor(exist_count/howmanynewfiles))
    print ('how_many_rows_per_file',how_many_rows_per_file)
    columns=generate_columnlist()
    template_row=cache_row(ws,1,columns)

    row_counter=0
    for i in range(1,howmanynewfiles+1):
        wb_new=Workbook()
        ws_new=wb_new.active

        # fill in data
        ws_new.append(template_row)
        for row in range(2,how_many_rows_per_file + 2):
            for col in range(1,len(columns)):
                ws_new.cell(row=row,column=col).value=ws.cell(row=row+row_counter,column=col).value
            # ws_new.append(cache_row(row,columns))
            # col_d = ws[row] # 0-indexing
            # for idx, cell in enumerate(col_d, 1):
            #     ws_new.cell(row=idx, column=4).value = cell.value #1-indexing
        # for row in range(row_counter,how_many_rows_per_file + row_counter):
        #     ws_new.append(ws[row])
        #     print(row)
        # ws_new[2:5]=ws[2:5]

        row_counter=row_counter+how_many_rows_per_file
        print (row_counter)
        print(filename+'-'+str(i))
        ws_new.title = ws_new_title
        wb_new.save(filename+'-'+str(i)+'.xlsx')

if __name__ == "__main__":
  main()
