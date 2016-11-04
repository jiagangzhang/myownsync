import sys
import openpyxl
import random

from openpyxl import Workbook
from openpyxl import load_workbook

wb1=load_workbook('product_list.xlsx')
ws1=wb1.active

wb2=load_workbook('dataheavy_inventory.xlsx')
ws2=wb2.active

# howManyNewInventory=3000

# count how many records exists, only count column-A, start from row-2
def count_exist_rows(worksheet):
    count=0
    for row in worksheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value and len(cell.value)>0:
                # print(cell.value)
                count+=1
    return count
    # print ('Exist:',count)

def delete_rows(start,howmanyrows,columns):
    for i in range(start,howmanyrows+2):
        for s in columns:
            ws2['%s%d'%(s,i)]=None


def main():
    delete_exist=0
    product_count=count_exist_rows(ws1)
    exist_inventory=count_exist_rows(ws2)

    args = sys.argv[1:]
    if not args:
      howManyNewInventory=product_count
    else:
        if args[0] == '--d':
            delete_exist=1
            del args[0:1]
            if not args:
                print ('Delete all rows!')
                howManyNewInventory=0
            else:
                try:
                    howManyNewInventory=int(args[0])
                    if howManyNewInventory<0:
                        print ('\n\tIdiot')
                        sys.exit(1)
                except:
                    print ("usage: [--d] [Number]. "+'Must specify a number')
                    sys.exit(1)
        else:
            try :
                howManyNewInventory=int(args[0])
                if howManyNewInventory<0:
                    print ('\n\tWrong format')
                    sys.exit(1)
            except:
                print ("usage: [--d] [Number]. "+'Must specify a number')
                sys.exit(1)

    if howManyNewInventory>product_count:
        print ('Not enough existing products to create inventory')
        sys.exit(1)

    if delete_exist:
        delete_rows(2,exist_inventory,['A','B','C'])
        exist_inventory=0

    if howManyNewInventory>0:
        for i in range(exist_inventory+2,exist_inventory+howManyNewInventory+2):
            ws2['A%d'%(i)]=ws1['B%d'%(i - exist_inventory)].value
            ws2['B%d'%(i)]=9999-random.randint(1,1000)
            ws2['C%d'%(i)]='N'

    print ('how many new:',howManyNewInventory)
    print ('how many now:',howManyNewInventory+exist_inventory)

    wb2.save('dataheavy_inventory.xlsx')

if __name__ == "__main__":
  main()
