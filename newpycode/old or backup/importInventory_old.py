import openpyxl
import random

from openpyxl import Workbook
from openpyxl import load_workbook

wb1=load_workbook('product_list.xlsx')
ws1=wb1.active

wb2=load_workbook('dataheavy_inventory.xlsx')
ws2=wb2.active


howmanyinventory=12
exist_product=0
exist_inventory=0
# for row in ws1.iter_rows(min_row=1, max_col=1):
#     for cell in row:
#         if cell.value and len(cell.value)>0:
#             # print(cell.value)
#             exist_count+=1
#
# print ('Exist Products:',exist_count)
# for i in range(1,100):
#     a=ws['A%d'%(i)]
#     if a.value and len(a.value)>0:
#         print(a.value)
#         count+=1
for row in ws2.iter_rows(min_row=2, max_col=1):
    for cell in row:
        if cell.value and len(cell.value)>0:
            # print(cell.value)
            exist_inventory+=1

print ('Exist Inventory:',exist_inventory)

if (exist_inventory)<=howmanyinventory:
    for i in range(2,howmanyinventory+2):
        ws2['A%d'%(i)]=ws1['B%d'%(i)].value
        ws2['B%d'%(i)]=9999-random.randint(1,1000)
        ws2['C%d'%(i)]='N'
else:
    for i in range(2,exist_inventory+2):
        if i<=(howmanyinventory+1):
            ws2['A%d'%(i)]=ws1['B%d'%(i)].value
            ws2['B%d'%(i)]=9999-random.randint(1,1000)
            ws2['C%d'%(i)]='N'
        else:
            ws2['A%d'%(i)]=None
            ws2['B%d'%(i)]=None
            ws2['C%d'%(i)]=None

print ('final i:',i)
print ('how many now:',howmanyinventory)



wb2.save('dataheavy_inventory.xlsx')
