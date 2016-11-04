import openpyxl
import random

from openpyxl import Workbook
from openpyxl import load_workbook

wb1=load_workbook('dataheavy_product_verify.xlsx')
ws1=wb1.active

wb2=load_workbook('dataheavy_inventory.xlsx')
ws2=wb2.active


howmanyinventory=3000
exist_product=0
exist_inventory=0
# for row in ws1.iter_rows(min_row=1, max_col=1):
#     for cell in row:
#         if cell.value and len(cell.value)>0:
#             # print(cell.value)
#             exist_count+=1
#
# print ('Exist Products:',exist_count)
# for i in range(1,100):
#     a=ws['A%d'%(i)]
#     if a.value and len(a.value)>0:
#         print(a.value)
#         count+=1
for row in ws2.iter_rows(min_row=2, max_col=1):
    for cell in row:
        if cell.value and len(cell.value)>0:
            # print(cell.value)
            exist_inventory+=1

print ('Exist Inventory:',exist_inventory)

if (exist_inventory)<=howmanyinventory:
    for i in range(2,howmanyinventory+2):
        ws2['A%d'%(i)]=ws1['B%d'%(i)].value
        ws2['B%d'%(i)]=9999-random.randint(1,1000)
        ws2['C%d'%(i)]='N'
else:
    for i in range(2,exist_inventory+2):
        if i<=(howmanyinventory+1):
            ws2['A%d'%(i)]=ws1['B%d'%(i)].value
            ws2['B%d'%(i)]=9999-random.randint(1,1000)
            ws2['C%d'%(i)]='N'
        else:
            ws2['A%d'%(i)]=None
            ws2['B%d'%(i)]=None
            ws2['C%d'%(i)]=None

    # randstr=''.join(random.choice('0123456789'+ascii_uppercase) for a in range(11))
    # for k,v in {'-PINK-S':123,'-PINK-M':456,'-PINK-L':789,'-PINK-XL':101}.items():
    #     ws['A%d'%(i)]='qap-dh-'+randstr
    #     ws['B%d'%(i)]='qap-dh-'+randstr+k
    #     ws['C%d'%(i)]='h-'+randstr+str(v)
    #     ws['D%d'%(i)]=ws['D2'].value
    #     ws['E%d'%(i)]=0
    #     ws['F%d'%(i)]='CN'
    #     ws['H%d'%(i)]=0
    #     ws['I%d'%(i)]=ws['I2'].value
    #     ws['O%d'%(i)]=0
    #     ws['P%d'%(i)]=ws['P2'].value
    #     ws['Q%d'%(i)]=ws['Q2'].value
    #     ws['V%d'%(i)]=ws['V2'].value
    #     ws['W%d'%(i)]=ws['W2'].value
    #     ws['X%d'%(i)]=ws['X2'].value
    #     ws['Y%d'%(i)]=ws['Y2'].value
    #     ws['Z%d'%(i)]=ws['Z2'].value
    #     ws['AA%d'%(i)]=ws['AA2'].value
    #     ws['AB%d'%(i)]=ws['AB2'].value
    #     ws['AC%d'%(i)]=ws['AC2'].value
    #     ws['AD%d'%(i)]=ws['AC2'].value
    #     ws['AE%d'%(i)]=ws['AE2'].value
    #     ws['AF%d'%(i)]=ws['AF2'].value
    #     ws['AG%d'%(i)]=ws['AG2'].value
    #     ws['AH%d'%(i)]=ws['AH2'].value
    #     ws['AI%d'%(i)]=ws['AI2'].value
    #     ws['AJ%d'%(i)]=ws['AJ2'].value
    #     ws['AK%d'%(i)]=0
    #     ws['AL%d'%(i)]=1
    #     # ws['AM%d'%(i)]='Active'
    #     # print (ws['A%d'%(i)].value, ws['B%d'%(i)].value, ws['C%d'%(i)].value,ws['D%d'%(i)].value,ws['I%d'%(i)].value,ws['W%d'%(i)].value,ws['AJ%d'%(i)].value)
    #     i+=1
        # print ('i in dic loop is: ',i)
    # print ('i in outer loop is:',i)

print ('final i:',i)
print ('how many now:',howmanyinventory)



wb2.save('dataheavy_inventory.xlsx')
