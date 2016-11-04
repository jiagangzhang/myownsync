import sys
import openpyxl
import random

from string import ascii_uppercase
import itertools

from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")

wb=load_workbook('dataheavy_product_verify.xlsx')
ws=wb.active

wb2=load_workbook('dataheavy_product_template.xlsx')
ws2=wb2.active
# count how many records exists, only count column-A, start from row-2
def count_exist_rows():
    count=0
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value and len(cell.value)>0:
                # print(cell.value)
                count+=1
    return count
    # print ('Exist:',count)

# generate a list of A to Z......Z
def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)
        size +=1

#  create column from A,B,C to AL,AM... default is 39 columns(AM)
def generate_columnlist(opt=39):
    columnlist=[]
    for s in itertools.islice(iter_all_strings(), opt):
        columnlist.append(s)
    return columnlist

# cache the format of excel of the first row if delete all existing records
def cache_first_row_format():
    cache=[]
    for s in generate_columnlist():
        cache.append(ws2['%s2'%s].value)
    return cache

# delete existing records, default is start=2, delete from product 1 to last
def delete_rows(start,howmanyrows,columns):
    for i in range(start,howmanyrows+2):
        for s in columns:
            ws['%s%d'%(s,i)]=None


def main():
    delete_exist=0

    args = sys.argv[1:]
    if not args:
      print ("usage: [--d] Number");
      sys.exit(1)

    if args[0] == '--d':
      delete_exist=1
      del args[0:1]

    if not args:
      print ("usage: [--d] Number");
      sys.exit(1)
    try:
        howmanynewproducts=int(args[0])
        if howmanynewproducts<0:
            print ('\n\tIdiot')
            sys.exit(1)
        elif howmanynewproducts==0: Print ('Delete all records!!')
    except:
        print ("usage: [--d] Number. "+'Must specify a number')
        sys.exit(1)

    exist_count=count_exist_rows()
    print ('\n\tExist:',exist_count)

    column_list=generate_columnlist()
    template_row=cache_first_row_format()

    if delete_exist:
        delete_rows(2,exist_count,column_list)
        exist_count=0

    new_rows=1
    for i in range(exist_count+2,exist_count+howmanynewproducts+2,4):
        randstr=''.join(random.choice('0123456789'+ascii_uppercase) for a in range(11))
        for k,v in {'-PINK-S':['123','S'],'-PINK-M':['456','M'],'-PINK-L':['789','L'],'-PINK-XL':['101','XL']}.items():
            ws['A%d'%(i)]='qap-dh-'+randstr
            ws['B%d'%(i)]='qap-dh-'+randstr+k
            ws['C%d'%(i)]='h-'+randstr+v[0]
            ws['D%d'%(i)]=template_row[3]
            ws['E%d'%(i)]=0
            ws['F%d'%(i)]='CN'
            ws['H%d'%(i)]=0
            ws['I%d'%(i)]=template_row[8]
            ws['O%d'%(i)]=0
            ws['P%d'%(i)]=template_row[15]
            ws['Q%d'%(i)]=template_row[16]
            ws['V%d'%(i)]=template_row[21]
            ws['W%d'%(i)]=v[1]
            ws['X%d'%(i)]=template_row[23]
            ws['Y%d'%(i)]=template_row[24]
            ws['Z%d'%(i)]=template_row[25]
            ws['AA%d'%(i)]=template_row[26]
            ws['AB%d'%(i)]=template_row[27]
            ws['AC%d'%(i)]=template_row[28]
            ws['AD%d'%(i)]=template_row[29]
            ws['AE%d'%(i)]=template_row[30]
            ws['AF%d'%(i)]=template_row[31]
            ws['AG%d'%(i)]=template_row[32]
            ws['AH%d'%(i)]=template_row[33]
            ws['AI%d'%(i)]=template_row[34]
            ws['AJ%d'%(i)]=template_row[35]
            ws['AK%d'%(i)]=0
            ws['AL%d'%(i)]=1
            # ws['AM%d'%(i)]='Active'
            # print (ws['A%d'%(i)].value, ws['B%d'%(i)].value, ws['C%d'%(i)].value,ws['D%d'%(i)].value,ws['I%d'%(i)].value,ws['W%d'%(i)].value,ws['AJ%d'%(i)].value)
            i+=1
            new_rows+=1
            # print ('i in dic loop is: ',i)
    # print ('i in outer loop is:',i)
    print ('\tNew Rows Created:',new_rows-1)
    wb.save('dataheavy_product_verify.xlsx')

if __name__ == "__main__":
  main()
