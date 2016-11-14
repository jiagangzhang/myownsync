import sys
import openpyxl
import random
import itertools
from string import ascii_uppercase

from openpyxl import Workbook
from openpyxl import load_workbook

wb1=load_workbook('confirmed_orders.xlsx')
ws1=wb1.active

wb2=load_workbook('Import_Shipment.xlsx')
ws2=wb2.active

# howManyNewShipment=3000

# generate a list of A to Z......Z
def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)
        size +=1

#  create column from A,B,C to AL,AM... default is 39 columns(AM)
def generate_columnlist(opt=39):
    columnlist=[]
    for s in itertools.islice(iter_all_strings(), opt):
        columnlist.append(s)
    return columnlist

# count how many records exists, only count column-A, start from row-2
def count_exist_rows(worksheet):
    count=0
    for row in worksheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value and len(cell.value)>0:
                # print(cell.value)
                count+=1
    return count
    # print ('Exist:',count)

def delete_rows(start,howmanyrows,columns):
    for i in range(start,howmanyrows+2):
        for s in columns:
            ws2['%s%d'%(s,i)]=None


def main():
    delete_exist=1
    order_count=count_exist_rows(ws1)
    exist_shipment=count_exist_rows(ws2)

    args = sys.argv[1:]
    if not args:
        print ('\n\t\tWill delete shipment xlsx and fill with new shipment records!!\n\n')
        howManyNewShipment=order_count
    else:
        if args[0] == '--append':
            delete_exist=0
            del args[0:1]
            if not args:
                print ('\n\n\t\tWill append all orders to the existing shipment list!!\n')
                howManyNewShipment=order_count
            else:
                try:
                    howManyNewShipment=int(args[0])
                    if howManyNewShipment<0:
                        print ('\n\tIdiot')
                        sys.exit(1)
                except:
                    print ("usage: [--append] [Number]. "+'Must specify a number')
                    quit()
        else:
            try :
                howManyNewShipment=int(args[0])
                if howManyNewShipment<0:
                    print ('\n\tWrong format')
                    sys.exit(1)
            except:
                print ("usage: [--append] [Number]. "+'Must specify a number')
                sys.exit(1)

    if howManyNewShipment>order_count:
        print ('Not enough existing orders to create shipment')
        sys.exit(1)

    if delete_exist:
        column_list=generate_columnlist(22)
        delete_rows(2,exist_shipment,column_list)
        exist_shipment=0

    if howManyNewShipment>0:
        for i in range(exist_shipment+2,exist_shipment+howManyNewShipment+2):
            ws2['A%d'%(i)]=ws1['A%d'%(i - exist_shipment)].value
            # ws2['B%d'%(i)]=9999-random.randint(1,1000)
            ws2['C%d'%(i)]='shunfeng'
            ws2['D%d'%(i)]='sf-'+str(random.randint(100000,999999))
            ws2['E%d'%(i)]='SHIPPED'
            ws2['F%d'%(i)]=ws1['I%d'%(i - exist_shipment)].value
            ws2['G%d'%(i)]=ws1['J%d'%(i - exist_shipment)].value
            ws2['H%d'%(i)]=ws1['K%d'%(i - exist_shipment)].value
            ws2['I%d'%(i)]=ws1['L%d'%(i - exist_shipment)].value
            ws2['J%d'%(i)]=ws1['M%d'%(i - exist_shipment)].value
            ws2['K%d'%(i)]=ws1['N%d'%(i - exist_shipment)].value
            ws2['N%d'%(i)]=ws1['Q%d'%(i - exist_shipment)].value
            ws2['T%d'%(i)]=ws1['AC%d'%(i - exist_shipment)].value
            ws2['U%d'%(i)]=int(ws1['AJ%d'%(i - exist_shipment)].value)

    print ('\t\thow many new:',howManyNewShipment)
    print ('\t\thow many now:',howManyNewShipment+exist_shipment,'\n')

    wb2.save('Import_Shipment.xlsx')

if __name__ == "__main__":
  main()
