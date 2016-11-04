#! /usr/bin/env python3
import openpyxl
import random

from string import ascii_uppercase
import itertools

from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")

wb=load_workbook('dataheavy_product_verify.xlsx')
ws=wb.active


howmanynewproducts=3000
exist_count=0
delete_exist=1
columnlist=[]
first_row_cache=[]

for row in ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        if cell.value and len(cell.value)>0:
            # print(cell.value)
            exist_count+=1
print ('Exist:',exist_count)

#  create column from A,B,C to AL,AM
def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)
        size +=1

for s in itertools.islice(iter_all_strings(), 39):
    columnlist.append(s)
for s in columnlist:
    first_row_cache.append(ws['%s2'%s].value)

def delete_rows(start,howmanyrows,columns):
    for i in range(start,howmanyrows):
        for s in columns:
            ws['%s%d'%(s,i)]=None


if delete_exist:
    delete_rows(2,exist_count,columnlist)
    exist_count=1


for i in range(exist_count+1,exist_count+howmanynewproducts+1,4):
    randstr=''.join(random.choice('0123456789'+ascii_uppercase) for a in range(11))
    for k,v in {'-PINK-S':['123','S'],'-PINK-M':['456','M'],'-PINK-L':['789','L'],'-PINK-XL':['101','XL']}.items():
        ws['A%d'%(i)]='qap-dh-'+randstr
        ws['B%d'%(i)]='qap-dh-'+randstr+k
        ws['C%d'%(i)]='h-'+randstr+v[0]
        ws['D%d'%(i)]=first_row_cache[3]
        ws['E%d'%(i)]=0
        ws['F%d'%(i)]='CN'
        ws['H%d'%(i)]=0
        ws['I%d'%(i)]=first_row_cache[8]
        ws['O%d'%(i)]=0
        ws['P%d'%(i)]=first_row_cache[15]
        ws['Q%d'%(i)]=first_row_cache[16]
        ws['V%d'%(i)]=first_row_cache[21]
        ws['W%d'%(i)]=v[1]
        ws['X%d'%(i)]=first_row_cache[23]
        ws['Y%d'%(i)]=first_row_cache[24]
        ws['Z%d'%(i)]=first_row_cache[25]
        ws['AA%d'%(i)]=first_row_cache[26]
        ws['AB%d'%(i)]=first_row_cache[27]
        ws['AC%d'%(i)]=first_row_cache[28]
        ws['AD%d'%(i)]=first_row_cache[29]
        ws['AE%d'%(i)]=first_row_cache[30]
        ws['AF%d'%(i)]=first_row_cache[31]
        ws['AG%d'%(i)]=first_row_cache[32]
        ws['AH%d'%(i)]=first_row_cache[33]
        ws['AI%d'%(i)]=first_row_cache[34]
        ws['AJ%d'%(i)]=first_row_cache[35]
        ws['AK%d'%(i)]=0
        ws['AL%d'%(i)]=1
        # ws['AM%d'%(i)]='Active'
        # print (ws['A%d'%(i)].value, ws['B%d'%(i)].value, ws['C%d'%(i)].value,ws['D%d'%(i)].value,ws['I%d'%(i)].value,ws['W%d'%(i)].value,ws['AJ%d'%(i)].value)
        i+=1
        # print ('i in dic loop is: ',i)
    # print ('i in outer loop is:',i)

print ('final i:',i-1)

wb.save('dataheavy_product_verify.xlsx')
