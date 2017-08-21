import sys
import openpyxl
import random
import math
import time

from string import ascii_uppercase
import itertools

from openpyxl import Workbook
from openpyxl import load_workbook
# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")

new_file_name = 'Coupon_Import.xlsx'
template_file = 'Coupon_BatchUploadTemplate.xlsx'

# wb2=load_workbook('dataheavy_product_template.xlsx')
# ws2=wb2.active

user_list=[]
with open ('active_user.csv') as f:
    for uname in f:
        uname=uname.rstrip()
        if uname == 'UserName' or uname is None:
            continue
        else:
            user_list.append(uname)
active_user=len(user_list)
print ('\n\tTotal active user is',active_user)

# count how many records exists, only count column-A, start from row-2
def count_exist_rows():
    count=0
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value and len(cell.value)>0:
                # print(cell.value)
                count+=1
    return count
    # print ('Exist:',count)

# generate a list of A to Z......Z
def iter_all_strings():
    size = 1
    while True:
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)
        size +=1

#  create column from A,B,C to AL,AM... default is 39 columns(AM)
def generate_columnlist(opt=39):
    columnlist=[]
    for s in itertools.islice(iter_all_strings(), opt):
        columnlist.append(s)
    return columnlist

# cache the format of excel of the first row if delete all existing records
def cache_first_row_format():
    cache=[]
    for s in generate_columnlist():
        cache.append(ws2['%s2'%s].value)
    return cache

# delete existing records, default is start=2, delete from product 1 to last
def delete_rows(start,howmanyrows,columns):
    for i in range(start,howmanyrows+2):
        for s in columns:
            ws['%s%d'%(s,i)]=None


def main():
    args = sys.argv[1:]
    if not args:
        new_coupon = active_user
        print ("\tWill create %d new coupons for all active user"%(new_coupon));
    else:
        try:
            new_coupon=int(args[0])
            if new_coupon<=0:
                print ('\n\tWrong format\n\tusage: [Number]')
                sys.exit(1)
            elif new_coupon > active_user:
                print ('Error, new coupon larger than exsiting active user')
                sys.exit(1)
        except:
            print ("usage: python3 coupon_batch.py [Number] "+' Must specify an integer')
            sys.exit(1)

    wb_new=Workbook()
    ws_new=wb_new.active

    wb=load_workbook(filename=template_file,data_only = True)
    ws=wb.active
    # column_list=generate_columnlist()
    # template_row=cache_first_row_format()
    for col in range(1,15):
        ws_new.cell(row=1,column=col).value=ws.cell(row=1,column=col).value

    new_rows=1
    for i in range(2,new_coupon+2):
        # randstr=''.join(random.choice('0123456789'+ascii_uppercase) for a in range(11))
        # for k,v in {'-PINK-S':['123','S'],'-PINK-M':['456','M'],'-PINK-L':['789','L'],'-PINK-XL':['101','XL']}.items():
            # if i > exist_count+howmanynewproducts+1 : break
        ws_new['A%d'%(i)]='qa'+'美美券 满288立减'
        ws_new['B%d'%(i)]='qa'+'美美券 满288立减'
        ws_new['C%d'%(i)]='qa'+'美美券 满288立减'
        ws_new['D%d'%(i)]=random.randint(500,1000)
        ws_new['E%d'%(i)]=random.randint(500,1000)-random.randint(200,500)
        ws_new['H%d'%(i)]=user_list[i-2]
        ws_new['J%d'%(i)]=1
        ws_new['K%d'%(i)]=time.strftime('%Y-%m-%d %H:%M:%S')
        if time.localtime()[0] > 2017:
            ws_new['L%d'%(i)]='2018-12-31 23:59:59'
        else:
            ws_new['L%d'%(i)]='2017-12-31 23:59:59'
        ws_new['M%d'%(i)]=random.choice(['Y',''])

            # ws['AM%d'%(i)]='Active'
            # print (ws['A%d'%(i)].value, ws['B%d'%(i)].value, ws['C%d'%(i)].value,ws['D%d'%(i)].value,ws['I%d'%(i)].value,ws['W%d'%(i)].value,ws['AJ%d'%(i)].value)
            # i+=1
        new_rows+=1
            # print ('i in dic loop is: ',i)
    # print ('i in outer loop is:',i)
    print ('\tNew Rows Created:',new_rows-1)
    if (new_rows-1) != new_coupon:
        print ('\n\t!!! New coupon created does NOT equal to desired coupon number, please check !!!')
    ws_new.title = ws.title
    wb_new.save(new_file_name)

if __name__ == "__main__":
  main()
